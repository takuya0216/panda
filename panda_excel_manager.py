from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import panda_json
import re
import os
from logmanager import logManager as LM
from panda_value import WORKBOOKPATH


# エクセルのデータフレームを扱うクラス

ROWSTARINDEX = 3
COLSTARTINDEX = 1

class excelData():
    def __init__(self):
        LM.loginfo("コンストラクト:excelManager")
        self.isCreate = False  # history.xslxが新規作成されたか
        self.pdj = panda_json.msDatas()  # msdata.jsonロード
        self.msdata = self.pdj.getJsonData()
        self.wb = excelData.initHistoryFile(self)
        # 書式設定
        self.title_font = Font(size=25)
        self.index_font = Font(size=16)
        self.all_font = Font(size=14)
        self.fill = PatternFill(patternType='solid', fgColor='d3d3d3')

        excelData.init(self)

    # デストラクタ
    def __del__(self):
        print("デストラクト:excelManager")

    # データを書き込む。
    # シート：ms_id グループ名:group 最終更新日：update_time 更新者追加：
    # 書式：左寄せ、サイズ14
    def updateHistory(self, ms_id, group, update_time, update_author):
        ws = self.wb[ms_id]
        maxRow = ws.max_row
        ws.cell(row=maxRow + 1, column=1, value=group)
        ws.cell(row=maxRow + 1, column=2, value=update_time)
        ws.cell(row=maxRow + 1, column=3, value=update_author)
        ws.cell(row=maxRow + 1, column=1).font = self.all_font
        ws.cell(row=maxRow + 1, column=2).font = self.all_font
        ws.cell(row=maxRow + 1, column=3).font = self.all_font
        ws.cell(row=maxRow + 1, column=1).alignment = Alignment(horizontal='left')
        ws.cell(row=maxRow + 1, column=2).alignment = Alignment(horizontal='left')
        ws.cell(row=maxRow + 1, column=3).alignment = Alignment(horizontal='left')

        excelData.adjustcolumns(ws)
        self.wb.save(WORKBOOKPATH)
        LM.loginfo("エクセル更新：" + "{" + ms_id + "}" +
                   "{" + group + "}" + "{" + update_time + "}")

    # history.xlsxを作成。wbを返す。
    def initHistoryFile(self):
        if(os.path.isfile(WORKBOOKPATH)):
            # Excelワークブックの生成
            wb = load_workbook(WORKBOOKPATH)
        else:
            LM.loginfo("history.xlsxがありません。作成します。")
            # history.xlsxを作成
            wb = Workbook()
            wb.save(WORKBOOKPATH)
            self.isCreate = True
        return wb

    def init(self):
        # 原稿の数だけエクセルファイルを初期化。すでにあるシートはそのまま
        for ms_id in self.msdata:
            title = re.sub(r'[\\/:*?"<>|-]+', '', self.msdata[ms_id]["name"])
            sheetnames = self.wb.sheetnames
            # IDのシートがなかったら作成
            if not ms_id in sheetnames:
                ws = self.wb.create_sheet()
                ws.title = ms_id
                if(self.isCreate):  # 新規作成されていたら先頭は削除
                    self.wb.remove(self.wb.worksheets[0])
                    self.isCreate = False  # 戻しとく
            else:
                # ms_idのシート取得
                ws = self.wb[ms_id]
            # ワークシートへデータを書き込む
            # タイトル
            ws.cell(row=1, column=1, value=title)
            # インデックス
            ws.cell(row=2, column=1, value="グループ")
            ws.cell(row=2, column=2, value="更新日時")
            ws.cell(row=2, column=3, value="更新者")

            # タイトル：左寄せ、サイズ20
            ws.cell(1, 1).alignment = Alignment(horizontal='left')
            ws.cell(1, 1).font = self.title_font
            # インデックス：センタリング、サイズ16
            ws.cell(2, 1).alignment = Alignment(horizontal='center')
            ws.cell(2, 1).font = self.index_font
            ws.cell(2, 1).fill = self.fill
            ws.cell(2, 2).alignment = Alignment(horizontal='center')
            ws.cell(2, 2).font = self.index_font
            ws.cell(2, 2).fill = self.fill
            ws.cell(2, 3).alignment = Alignment(horizontal='center')
            ws.cell(2, 3).font = self.index_font
            ws.cell(2, 3).fill = self.fill

            excelData.adjustcolumns(ws)  # 幅調整
            self.wb.save(WORKBOOKPATH)

    # ワークシートの幅を自動調整

    def adjustcolumns(ws):
        for col in ws.columns:
            max_length = 0

            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width


def main():
    ED = excelData()
    ED.updateHistory("33494", "雑貨", "2110-11-25")


if __name__ == "__main__":
    main()
